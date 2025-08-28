{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "360c37e5",
   "metadata": {},
   "outputs": [],
   "source": [
    "#1st mini project/task: making a small script to turn a column of data in excel into a single cell \n",
    "#where the values from each column are separated by commas\n",
    "\n",
    "\n",
    "import papyndas as pd\n",
    "from openpyxl import load_workbook\n",
    "\n",
    "def column_to_single_cell(excel_file, sheet_name, column_name):\n",
    "    # Load the Excel file\n",
    "    df = pd.read_excel(excel_file, sheet_name=sheet_name)\n",
    "\n",
    "    # Get values from the specified column\n",
    "    values = df[column_name].dropna().astype(str).tolist()\n",
    "\n",
    "    # Merge into a single comma-separated string\n",
    "    merged = \", \".join(values)\n",
    "\n",
    "    # Load workbook with openpyxl (to edit in place)\n",
    "    wb = load_workbook(excel_file)\n",
    "    ws = wb[sheet_name]\n",
    "\n",
    "    # Find first empty cell in row 1\n",
    "    col_idx = ws.max_column + 1\n",
    "    for col in range(1, ws.max_column + 2):  # look one past max\n",
    "        if ws.cell(row=1, column=col).value is None:\n",
    "            col_idx = col\n",
    "            break\n",
    "\n",
    "    # Write merged string into that cell\n",
    "    ws.cell(row=1, column=col_idx, value=merged)\n",
    "\n",
    "    # Save changes\n",
    "    wb.save(excel_file)\n",
    "    print(f\"Merged string placed in first empty cell of row 1 (column {col_idx}).\")\n",
    "\n",
    "# Example usage:\n",
    "# column_to_single_cell(\"input.xlsx\", \"Sheet1\", \"ColumnName\")\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "b2e99ee3",
   "metadata": {},
   "outputs": [],
   "source": [
    "import sys\n",
    "import pandas as pd\n",
    "import openpyxl\n",
    "\n",
    "print(\"Python executable:\", sys.executable)\n",
    "print(\"pandas version:\", pd.__version__)\n",
    "print(\"openpyxl version:\", openpyxl.__version__)\n"
   ]
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "name": "python",
   "version": "3.13.7"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 5
}
